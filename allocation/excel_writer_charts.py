from openpyxl.chart import PieChart, Reference, BarChart
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.legend import Legend
from openpyxl.chart.marker import DataPoint
from openpyxl.reader.excel import load_workbook


class ExcelWriterCharts:
    def __init__(self, results_file_path, preferences_met_file_path):
        self.results_file_path = results_file_path
        self.preferences_met_file_path = preferences_met_file_path


    def add_courses_sat_bar_chart(self):
        try:
            wb = load_workbook(self.results_file_path)
            ws = wb.active

            wb.create_sheet("Course Sat Bar Chart")
            bar_ws = wb["Course Sat Bar Chart"]

            course_satisfaction = {}

            last_row = ws.max_row
            for row in range(2, last_row + 1):
                course_name = ws.cell(row = row, column = 1).value  # 1: "Course Name"
                student_preference = ws.cell(row = row, column = 7).value  # 7: "Preference"

                if course_name not in course_satisfaction:
                    course_satisfaction[course_name] = {"count": 0, "1st_preference_count": 0}

                course_satisfaction[course_name]["count"] += 1
                if student_preference == 1:
                    course_satisfaction[course_name]["1st_preference_count"] += 1

            chart_data = []
            for course_name, data in course_satisfaction.items():
                satisfaction_ratio = round(data["1st_preference_count"] / data["count"] * 100, 2)
                chart_data.append([course_name, satisfaction_ratio])

            bar_ws.append(["Course", "1st Preference Satisfaction Ratio (%)"]) # headers
            for row in chart_data:
                bar_ws.append(row)


            chart = BarChart()
            data = Reference(bar_ws, min_col=2, min_row=1, max_col=2, max_row=len(chart_data) + 1) # "Satisfaction Ratio"
            categories = Reference(bar_ws, min_col=1, min_row=2, max_row=len(chart_data) + 1) # "Course"

            chart.add_data(data, titles_from_data = True)
            chart.set_categories(categories)
            chart.title = "1st Preference Satisfaction Ratio (%)"
            chart.x_axis.title = "Courses"
            chart.y_axis.title = "Satisfaction Ratio"
            chart.width = 15
            chart.height = 15

            series = chart.series[0]

            series.dLbls = DataLabelList()
            series.dLbls.showVal = True
            series.dLbls.numFmt = "0.00"
            series.dLbls.showCatName = False
            series.dLbls.showLegendKey = False
            series.dLbls.showSerName = False
            series.dLbls.showPercent = False

            bar_ws.add_chart(chart, "E5")

            wb.save(self.results_file_path)
            print(f"1st preference satisfaction bar chart added to {self.results_file_path}")

        except Exception as e:
            print("An error occurred while adding the bar chart. \n", e)


    def add_preferences_met_pie_chart(self, preferences_met_ratios):
        try:
            wb = load_workbook(self.preferences_met_file_path)
            ws = wb.active

            if preferences_met_ratios:
                avg_preferences_met = round(sum(preferences_met_ratios) / len(preferences_met_ratios), 2)
            else:
                avg_preferences_met = 0

            avg_preferences_not_met = round(100 - avg_preferences_met, 2)

            ws.cell(row=2, column=11).value = "Met"
            ws.cell(row=2, column=10).value = avg_preferences_met
            ws.cell(row=3, column=11).value = "Not Met"
            ws.cell(row=3, column=10).value = avg_preferences_not_met

            pie = PieChart()

            data = Reference(ws, min_col=10, min_row=2, max_row=3)
            labels = Reference(ws, min_col=11, min_row=2, max_row=3)

            pie.add_data(data, titles_from_data = False)
            pie.set_categories(labels)
            pie.title = "Average Student Preferences (%)"
            pie.height = 10
            pie.width = 10

            pie.legend = Legend()
            pie.legend.position = "b"

            series = pie.series[0]

            series.dLbls = DataLabelList()
            series.dLbls.showVal = True
            series.dLbls.numFmt = "0.00"
            series.dLbls.showCatName = False
            series.dLbls.showLegendKey = False
            series.dLbls.showSerName = False
            series.dLbls.showPercent = False

            # green for preferences met
            met_slice = DataPoint(idx=0)
            met_slice.graphicalProperties.solidFill = "00B050"

            # red for preferences not met
            not_met_slice = DataPoint(idx=1)
            not_met_slice.graphicalProperties.solidFill = "FF0000"

            series.dPt = [met_slice, not_met_slice]

            ws.add_chart(pie, "J10")

            wb.save(self.preferences_met_file_path)

            print(f"Average student preferences pie chart added to {self.preferences_met_file_path}")

        except Exception as e:
            print("An error occurred while adding the pie chart. \n", e)




