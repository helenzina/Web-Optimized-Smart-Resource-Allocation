from openpyxl.chart import PieChart, Reference, BarChart
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.legend import Legend
from openpyxl.chart.marker import DataPoint
from openpyxl.utils import get_column_letter

from openpyxl.reader.excel import load_workbook


class ExcelWriterCharts:
    def __init__(self, results_file_path):
        self.results_file_path = results_file_path


    def add_top_6_preferences_sat_pie_chart(self, top_6_preferences_satisfaction_ratios):
        try:
            wb = load_workbook(self.results_file_path)
            ws = wb["Preferences Satisfaction"]

            if top_6_preferences_satisfaction_ratios:
                avg_preferences_met = round(
                    sum(top_6_preferences_satisfaction_ratios) / len(top_6_preferences_satisfaction_ratios), 2
                )
            else:
                avg_preferences_met = 0

            avg_preferences_not_met = round(100 - avg_preferences_met, 2)

            ws.cell(row=2, column=11).value = "Met"
            ws.cell(row=2, column=10).value = avg_preferences_met
            ws.cell(row=3, column=11).value = "Not Met"
            ws.cell(row=3, column=10).value = avg_preferences_not_met

            pie = PieChart()

            data = Reference(ws, min_col=10, min_row=2, max_row=3)
            labels = Reference(ws, min_col=11, min_row=2, max_row=3)

            pie.add_data(data, titles_from_data = False)
            pie.set_categories(labels)
            pie.title = "Average Student Top 6 Preferences (%)"
            pie.height = 10
            pie.width = 10

            pie.legend = Legend()
            pie.legend.position = "b"

            series = pie.series[0]

            series.dLbls = DataLabelList()
            series.dLbls.showVal = True
            series.dLbls.numFmt = "0.00"
            series.dLbls.showCatName = False
            series.dLbls.showLegendKey = False
            series.dLbls.showSerName = False
            series.dLbls.showPercent = False

            # green for preferences met
            met_slice = DataPoint(idx=0)
            met_slice.graphicalProperties.solidFill = "00B050"

            # red for preferences not met
            not_met_slice = DataPoint(idx=1)
            not_met_slice.graphicalProperties.solidFill = "FF0000"

            series.dPt = [met_slice, not_met_slice]

            ws.add_chart(pie, "L10")

            wb.save(self.results_file_path)

            print(f"Average student top 6 preferences pie chart added to {self.results_file_path}")

        except Exception as e:
            print("An error occurred while adding the pie chart. \n", e)

        return avg_preferences_met

    def add_courses_sat_bar_charts(self):
        """Adds bar charts for each preference-course pair."""
        try:
            wb = load_workbook(self.results_file_path)
            ws = wb.active

            if "Course Sat" not in wb.sheetnames:
                wb.create_sheet("Course Sat")
            bar_ws = wb["Course Sat"]

            course_satisfaction = {}

            last_row = ws.max_row
            for row in range(2, last_row + 1):
                course_name = ws.cell(row=row, column=1).value  # 1: "Course Name"
                student_preference = ws.cell(row=row, column=7).value  # 7: "Preference"
                
                if course_name not in course_satisfaction:
                    course_satisfaction[course_name] = {"count": 0}
                
                course_satisfaction[course_name]["count"] += 1
                
                if f"{student_preference}_preference_count" not in course_satisfaction[course_name]:
                    course_satisfaction[course_name][f"{student_preference}_preference_count"] = 0
                
                course_satisfaction[course_name][f"{student_preference}_preference_count"] += 1

            for p in range(len(course_satisfaction) + 1):
                for course_name in course_satisfaction:
                    if f"{p}_preference_count" not in course_satisfaction[course_name]:
                        course_satisfaction[course_name][f"{p}_preference_count"] = 0                
            
            headers = ["Course"]
            for p in range(1, len(course_satisfaction) + 1):
                headers.append(f"{p} Preference Satisfaction Ratio (%)")
            bar_ws.append(headers)

            chart_data = []
            for course_name, data in course_satisfaction.items():
                row_data = [course_name]
                for p in range(1, len(course_satisfaction) + 1):
                    satisfaction_ratio = (
                        round(data[f"{p}_preference_count"] / data["count"] * 100, 2) if data["count"] > 0 else 0
                    )
                    row_data.append(satisfaction_ratio)
                bar_ws.append(row_data)
                chart_data.append(row_data)

            for p in range(1, len(course_satisfaction) + 1):
                chart = BarChart()

                data = Reference(bar_ws, min_col=p+1, min_row=1, max_col=p+1, max_row=len(chart_data)+1)
                categories = Reference(bar_ws, min_col=1, min_row=2, max_row=len(chart_data)+1)

                chart.add_data(data, titles_from_data=True)
                chart.set_categories(categories)
                chart.title = f"{p} Preference Satisfaction Ratio (%)"
                chart.x_axis.title = "Courses"
                chart.y_axis.title = "Satisfaction Ratio"
                chart.width = 9
                chart.height = 15

                series = chart.series[0]
                series.dLbls = DataLabelList()
                series.dLbls.showVal = True
                series.dLbls.numFmt = "0.00"
                series.dLbls.showCatName = False
                series.dLbls.showLegendKey = False
                series.dLbls.showSerName = False
                series.dLbls.showPercent = False

                # add the chart below the data
                chart_row = bar_ws.max_row + 3 
                col_number = 1 + (p-1) * 2
                col_letter = get_column_letter(col_number)
                bar_ws.add_chart(chart, f"{col_letter}{chart_row}")
                        
            wb.save(self.results_file_path)
            print(f"Preferences satisfaction bar charts are added to {self.results_file_path}")

        except Exception as e:
            print("An error occurred while adding preferences satisfaction bar charts. \n", e)

